from __future__ import annotations

import logging
from typing import Iterable

from openpyxl import load_workbook

from .models import AppConfig, SatelliteRequest


def read_satellite_requests(
    config: AppConfig,
    logger: logging.Logger,
) -> list[SatelliteRequest]:
    workbook = load_workbook(
        filename=config.input_excel_path,
        read_only=True,
        data_only=True,
    )
    try:
        if config.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{config.sheet_name}' was not found in {config.input_excel_path.name}."
            )

        worksheet = workbook[config.sheet_name]
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if header_row is None:
            raise ValueError("The worksheet does not contain a header row.")

        header_map = {
            _normalize_header_name(value): index
            for index, value in enumerate(header_row)
        }
        required_headers = [config.sat_name_header, config.norad_id_header]
        missing_headers = [
            header
            for header in required_headers
            if _normalize_header_name(header) not in header_map
        ]
        if missing_headers:
            raise ValueError(
                "Missing required headers: " + ", ".join(missing_headers)
            )

        sat_name_index = header_map[_normalize_header_name(config.sat_name_header)]
        norad_id_index = header_map[_normalize_header_name(config.norad_id_header)]

        requests: list[SatelliteRequest] = []
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            sat_name = _clean_text(_safe_get(row_values, sat_name_index))
            norad_id = _normalize_norad_id(_safe_get(row_values, norad_id_index))

            if not sat_name and not norad_id:
                continue

            if not sat_name or not norad_id:
                logger.warning(
                    "Skipping row %s because SAT_Name or NORAD ID is empty.",
                    row_number,
                )
                continue

            requests.append(
                SatelliteRequest(
                    row_number=row_number,
                    sat_name=sat_name,
                    norad_id=norad_id,
                )
            )

        return requests
    finally:
        workbook.close()


def _safe_get(values: Iterable[object], index: int) -> object:
    values_list = list(values)
    if index >= len(values_list):
        return ""
    return values_list[index]


def _normalize_header_name(value: object) -> str:
    return _clean_text(value)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    return text


def _normalize_norad_id(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ""

    text = _clean_text(value)
    digits_only = text.replace(",", "")
    if digits_only.isdigit():
        return str(int(digits_only))
    return ""
